from openpyxl import load_workbook
from PIL import Image, ImageDraw

# Load the workbook
wb = load_workbook('template_Cyrus Rugs.xlsx')

# Select the first worksheet
ws = wb.worksheets[0]

# Get the dimensions of the worksheet
max_row = ws.max_row
max_col = ws.max_column

# Create a new image with a white background
img = Image.new('RGB', (max_col * 100, max_row * 30), color='white')
draw = ImageDraw.Draw(img)

# Draw the data from the worksheet onto the image
for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is not None:
            draw.text((col * 100 - 90, row * 30 - 20), str(cell_value), fill='black')

# Save the image as a PNG file
img.save('output.png')